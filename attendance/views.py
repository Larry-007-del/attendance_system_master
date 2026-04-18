from django.shortcuts import get_object_or_404
from rest_framework import viewsets, generics, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAdminUser, BasePermission
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.throttling import ScopedRateThrottle
from django.contrib.auth import authenticate, logout
from django.utils import timezone
from django.http import HttpResponse
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema, OpenApiResponse, OpenApiExample
import csv
from openpyxl import Workbook
from collections import defaultdict
from enum import Enum

from .models import Lecturer, Student, Course, Attendance, AttendanceToken, AttendanceStudent
from .serializers import (
    LecturerSerializer,
    StudentSerializer,
    CourseSerializer,
    AttendanceSerializer,
    AttendanceTokenSerializer,
    LogoutSerializer,
    SubmitLocationSerializer,
    APIErrorSerializer,
    StudentLoginRequestSerializer,
    StudentLoginSuccessSerializer,
    StaffLoginRequestSerializer,
    StaffLoginSuccessSerializer,
    AttendanceMarkedSerializer,
    LecturerLocationResponseSerializer,
    LecturerLocationRequestSerializer,
)
from .error_codes import APIErrorCode


def api_error(message, code, http_status=status.HTTP_400_BAD_REQUEST, details=None):
    if isinstance(code, Enum):
        code = code.value
    payload = {
        'error': message,
        'code': code,
    }
    if details is not None:
        payload['details'] = details
    return Response(payload, status=http_status)


# ==================== Custom Permissions ====================

class IsStaffOrAdmin(BasePermission):
    """Allow access to lecturers and superusers only."""
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return request.user.is_superuser or hasattr(request.user, 'lecturer')


class IsAdminOrReadOnly(BasePermission):
    """Allow read access to authenticated staff; write access to admins only."""
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in ('GET', 'HEAD', 'OPTIONS'):
            return request.user.is_superuser or hasattr(request.user, 'lecturer')
        return request.user.is_superuser

# Lecturer ViewSet
class LecturerViewSet(viewsets.ModelViewSet):
    queryset = Lecturer.objects.select_related('user').all().order_by('name')
    serializer_class = LecturerSerializer
    permission_classes = [IsAdminOrReadOnly]

    @action(detail=False, methods=['get'], url_path='my-courses')
    def my_courses(self, request):
        lecturer = get_object_or_404(Lecturer, user=request.user)
        courses = Course.objects.select_related('lecturer', 'lecturer__user').prefetch_related('students', 'students__user').filter(lecturer=lecturer)
        serializer = CourseSerializer(courses, many=True)
        return Response(serializer.data)

# Student ViewSet
class StudentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        """Students can read; only admins can create/update/delete."""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminOrReadOnly()]
        return super().get_permissions()

    def get_queryset(self):
        user = self.request.user
        # Lecturers can see all students
        if hasattr(user, 'lecturer'):
            return Student.objects.select_related('user').all().order_by('name')
        # Students can only see themselves
        elif hasattr(user, 'student'):
            return Student.objects.select_related('user').filter(user=user).order_by('name')
        return Student.objects.none()

# Course ViewSet
class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.select_related('lecturer', 'lecturer__user').prefetch_related('students', 'students__user').all().order_by('name')
    serializer_class = CourseSerializer
    permission_classes = [IsStaffOrAdmin]

    @action(detail=True, methods=['post'])
    def generate_attendance_token(self, request, pk=None):
        course = self.get_object()
        
        # Ownership check: only the course's lecturer or an admin can generate tokens
        if not request.user.is_superuser:
            if not hasattr(request.user, 'lecturer') or course.lecturer != request.user.lecturer:
                return api_error('You do not own this course.', APIErrorCode.COURSE_FORBIDDEN, status.HTTP_403_FORBIDDEN)
        
        token_value = request.data.get('token')
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')

        if not token_value or not latitude or not longitude:
            return api_error(
                'Token, latitude, and longitude are required.',
                APIErrorCode.MISSING_REQUIRED_FIELDS,
                status.HTTP_400_BAD_REQUEST,
                {'required': ['token', 'latitude', 'longitude']}
            )

        try:
            latitude = float(latitude)
            longitude = float(longitude)
        except (ValueError, TypeError):
            return api_error('Invalid latitude or longitude.', APIErrorCode.INVALID_GPS_COORDINATES, status.HTTP_400_BAD_REQUEST)

        # CRITICAL FIX: Create/Update the Attendance record immediately so students can find it
        # We use get_or_create to prevent duplicates if the button is clicked twice
        attendance, created = Attendance.objects.get_or_create(
            course=course,
            date=timezone.localdate(),
            defaults={
                'lecturer_latitude': latitude,
                'lecturer_longitude': longitude,
                'is_active': True,
                'created_by': request.user
            }
        )

        # If it already existed, update the location and ensure it is active
        if not created:
            attendance.lecturer_latitude = latitude
            attendance.lecturer_longitude = longitude
            attendance.is_active = True
            attendance.save()

        # Now create the token — expiry matches the attendance session duration
        AttendanceToken.objects.create(
            course=course,
            token=token_value,
            generated_at=timezone.now(),
            expires_at=timezone.now() + timezone.timedelta(hours=attendance.duration_hours),
            is_active=True
        )

        # Update the lecturer's stored location
        course.lecturer.latitude = latitude
        course.lecturer.longitude = longitude
        course.lecturer.save(update_fields=['latitude', 'longitude'])

        # Send notifications to students (asynchronously)
        from .tasks import send_attendance_started_notifications
        send_attendance_started_notifications.delay(attendance.id, token_value)

        # Schedule expiration reminder (15 minutes before expiry)
        from .tasks import schedule_attendance_expiration_reminder
        schedule_attendance_expiration_reminder(attendance, token_value)

        return Response({'status': 'Token generated and session started', 'token': token_value})

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def take_attendance(self, request):
        """Allow any authenticated student to submit attendance via token."""
        token = request.data.get('token')
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')

        if not token:
            return api_error('Token is required.', APIErrorCode.TOKEN_REQUIRED, status.HTTP_400_BAD_REQUEST)
            
        if not latitude or not longitude:
            return api_error('GPS coordinates are required.', APIErrorCode.MISSING_REQUIRED_FIELDS, status.HTTP_400_BAD_REQUEST)

        try:
            latitude = float(latitude)
            longitude = float(longitude)
        except (ValueError, TypeError):
            return api_error('Invalid GPS coordinates.', APIErrorCode.MISSING_REQUIRED_FIELDS, status.HTTP_400_BAD_REQUEST)

        try:
            attendance_token = AttendanceToken.objects.get(token=token, is_active=True)
            course = attendance_token.course
            student = get_object_or_404(Student, user=request.user)

            if not course.students.filter(pk=student.pk).exists():
                return api_error('Student is not enrolled in this course.', APIErrorCode.STUDENT_NOT_ENROLLED, status.HTTP_400_BAD_REQUEST)

            attendance = Attendance.objects.filter(
                course=course,
                date=timezone.localdate(),
                is_active=True
            ).first()
            
            if not attendance or not attendance.is_session_valid:
                return api_error('This attendance session has expired.', APIErrorCode.SESSION_EXPIRED, status.HTTP_400_BAD_REQUEST)

            if not attendance.is_within_radius(latitude, longitude):
                return api_error('You are outside the classroom boundary.', APIErrorCode.LOCATION_OUT_OF_RANGE, status.HTTP_400_BAD_REQUEST)
            
            from django.db import transaction
            with transaction.atomic():
                locked_attendance = Attendance.objects.select_for_update().get(pk=attendance.pk)
                AttendanceStudent.objects.update_or_create(
                    attendance=locked_attendance,
                    student=student,
                    defaults={
                        'latitude': latitude,
                        'longitude': longitude
                    }
                )
                locked_attendance.present_students.add(student)

            return Response({'message': 'Attendance recorded successfully.'}, status=status.HTTP_200_OK)

        except AttendanceToken.DoesNotExist:
            return api_error('Invalid or expired token.', APIErrorCode.INVALID_OR_EXPIRED_TOKEN, status.HTTP_400_BAD_REQUEST)
        
# Attendance ViewSet
class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.select_related('course', 'course__lecturer', 'course__lecturer__user').prefetch_related('present_students', 'present_students__user', 'course__students', 'course__students__user').all().order_by('-date')
    serializer_class = AttendanceSerializer
    permission_classes = [IsStaffOrAdmin]

    @action(detail=False, methods=['get'])
    def generate_excel(self, request):
        attendance_id = request.query_params.get('attendance_id')

        if not attendance_id:
            return api_error('attendance_id parameter is required.', APIErrorCode.ATTENDANCE_ID_REQUIRED, status.HTTP_400_BAD_REQUEST)

        attendance = get_object_or_404(Attendance, id=attendance_id)

        # Create an Excel workbook and add a worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance Report"

        # Add header row
        worksheet.append(['Student ID', 'Student Name', 'Date of Attendance', 'Status'])

        # Style headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill

        # Collect present students (single query)
        present_set = set(attendance.present_students.all())
        present_students = [(s.student_id, s.name) for s in present_set]

        # Collect missed students
        course_students = list(attendance.course.students.all())
        missed_students = [(s.student_id, s.name) for s in course_students if s not in present_set]

        # Write present students
        for student_id, student_name in sorted(present_students):
            worksheet.append([student_id, student_name, attendance.date, 'Present'])

        # Write absent students
        for student_id, student_name in sorted(missed_students):
            worksheet.append([student_id, student_name, attendance.date, 'Absent'])

        # Auto-adjust column widths
        worksheet.column_dimensions['B'].width = 25  # Name
        worksheet.column_dimensions['C'].width = 18  # Date

        # Create an HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="attendance_{attendance_id}.xlsx"'

        workbook.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export_csv/(?P<course_id>[^/.]+)')
    def export_attendance_csv(self, request, course_id=None):
        """Export attendance as CSV for a specific course using StreamingHttpResponse"""
        from django.http import StreamingHttpResponse
        
        course = get_object_or_404(Course, id=course_id)

        # Ownership check
        if not request.user.is_superuser:
            if not hasattr(request.user, 'lecturer') or course.lecturer != request.user.lecturer:
                return api_error('You do not have permission to view attendance for this course.', APIErrorCode.COURSE_FORBIDDEN, status.HTTP_403_FORBIDDEN)
        
        class Echo:
            """An object that implements just the write method of the file-like interface."""
            def write(self, value):
                return value

        def iter_items():
            # Header Row
            yield ['Date', 'Student ID', 'Student Name', 'Status', 'Time Marked']
            
            # Data Rows — use iterator() to avoid loading all into memory
            attendance_students = AttendanceStudent.objects.filter(
                attendance__course=course
            ).select_related('student', 'student__user', 'attendance').order_by('attendance__date', 'student__name').iterator(chunk_size=1000)
            
            for record in attendance_students:
                yield [
                    record.attendance.date,
                    record.student.student_id,
                    record.student.name,
                    'Present',
                    record.marked_at.strftime("%H:%M:%S") if record.marked_at else '',
                ]

        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer)
        
        response = StreamingHttpResponse(
            (writer.writerow(row) for row in iter_items()),
            content_type="text/csv"
        )
        response['Content-Disposition'] = f'attachment; filename="{course.course_code}_attendance.csv"'
        return response

    @action(detail=False, methods=['post'], url_path='end_attendance')
    def end_attendance(self, request):
        course_id = request.data.get('course_id')
        if not course_id:
            return api_error('course_id is required.', APIErrorCode.COURSE_ID_REQUIRED, status.HTTP_400_BAD_REQUEST)
        
        try:
            # Retrieve the most recent attendance for the course
            attendance = Attendance.objects.filter(course_id=course_id, is_active=True).latest('date')
        except Attendance.DoesNotExist:
            return api_error('No active attendance found for the course.', APIErrorCode.ATTENDANCE_NOT_FOUND, status.HTTP_404_NOT_FOUND)

        # Ownership check: only course lecturer or admin can end a session
        if not request.user.is_superuser:
            if not hasattr(request.user, 'lecturer') or attendance.course.lecturer != request.user.lecturer:
                return api_error('You do not own this course.', APIErrorCode.COURSE_FORBIDDEN, status.HTTP_403_FORBIDDEN)

        attendance.is_active = False
        attendance.ended_at = timezone.now()
        attendance.save()
        
        # Deactivate any active tokens for this course
        AttendanceToken.objects.filter(course_id=course_id, is_active=True).update(is_active=False)
        
        # Send notifications to students who missed the session
        from .tasks import send_missed_attendance_notifications
        send_missed_attendance_notifications.delay(attendance.id)

        return Response({'status': 'Attendance session ended successfully'}, status=status.HTTP_200_OK)
    

# AttendanceToken ViewSet
class AttendanceTokenViewSet(viewsets.ModelViewSet):
    queryset = AttendanceToken.objects.select_related('course', 'course__lecturer', 'course__lecturer__user').all().order_by('-generated_at')
    serializer_class = AttendanceTokenSerializer
    permission_classes = [IsStaffOrAdmin]

# Student Enrolled Courses View
class StudentEnrolledCoursesView(generics.ListAPIView):
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        student = get_object_or_404(Student, user=user)
        return Course.objects.select_related('lecturer', 'lecturer__user').prefetch_related('students', 'students__user').filter(students=student).order_by('name')

# Custom Login Views
class StudentLoginView(APIView):
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = 'student_login'
    permission_classes = []
    authentication_classes = []

    @extend_schema(
        tags=['Auth'],
        summary='Student login',
        description='Authenticate a student and receive JWT access/refresh token pair.',
        request=StudentLoginRequestSerializer,
        responses={
            200: OpenApiResponse(response=StudentLoginSuccessSerializer, description='Student login successful.'),
            400: OpenApiResponse(response=APIErrorSerializer, description='Validation/authentication failure.'),
            429: OpenApiResponse(response=APIErrorSerializer, description='Login rate limit exceeded.'),
        },
        auth=[],
        examples=[
            OpenApiExample(
                'Student login error',
                value={'error': 'Invalid credentials', 'code': 'invalid_credentials'},
                response_only=True,
                status_codes=['400'],
            ),
        ],
    )
    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        student_id = request.data.get('student_id')

        user = authenticate(request, username=username, password=password)
        if user and hasattr(user, 'student'):
            student = user.student

            if student.student_id == student_id:
                refresh = RefreshToken.for_user(user)
                return Response({
                    'access': str(refresh.access_token),
                    'refresh': str(refresh),
                    'user_id': user.student.id,
                    'username': user.username,
                    'student_id': student.student_id
                })
            else:
                return api_error('Invalid student ID', APIErrorCode.INVALID_STUDENT_ID, status.HTTP_400_BAD_REQUEST)

        return api_error('Invalid credentials', APIErrorCode.INVALID_CREDENTIALS, status.HTTP_400_BAD_REQUEST)

class StaffLoginView(APIView):
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = 'staff_login'
    permission_classes = []
    authentication_classes = []

    @extend_schema(
        tags=['Auth'],
        summary='Staff login',
        description='Authenticate a lecturer/staff member and receive JWT access/refresh token pair.',
        request=StaffLoginRequestSerializer,
        responses={
            200: OpenApiResponse(response=StaffLoginSuccessSerializer, description='Staff login successful.'),
            400: OpenApiResponse(response=APIErrorSerializer, description='Validation/authentication failure.'),
            429: OpenApiResponse(response=APIErrorSerializer, description='Login rate limit exceeded.'),
        },
        auth=[],
        examples=[
            OpenApiExample(
                'Staff login error',
                value={'error': 'Invalid staff ID', 'code': 'invalid_staff_id'},
                response_only=True,
                status_codes=['400'],
            ),
        ],
    )
    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        staff_id = request.data.get('staff_id')

        user = authenticate(request, username=username, password=password)
        if user and hasattr(user, 'lecturer'):
            lecturer = user.lecturer
            if lecturer.staff_id == staff_id:
                refresh = RefreshToken.for_user(user)

                return Response({
                    'access': str(refresh.access_token),
                    'refresh': str(refresh),
                    'user_id': user.lecturer.id,
                    'username': user.username,
                    'staff_id': lecturer.staff_id
                })
            else:
                return api_error('Invalid staff ID', APIErrorCode.INVALID_STAFF_ID, status.HTTP_400_BAD_REQUEST)

        return api_error('Invalid credentials', APIErrorCode.INVALID_CREDENTIALS, status.HTTP_400_BAD_REQUEST)

# Logout View
class LogoutView(generics.GenericAPIView):
    serializer_class = LogoutSerializer
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=['Auth'],
        summary='Logout',
        description='Blacklist the refresh token and end the session.',
        request={'type': 'object', 'properties': {'refresh': {'type': 'string'}}, 'required': ['refresh']},
        responses={205: None, 400: OpenApiResponse(response=APIErrorSerializer, description='Invalid or missing refresh token.')},
    )
    def post(self, request, *args, **kwargs):
        refresh_token = request.data.get('refresh')
        if not refresh_token:
            return api_error('Refresh token is required', APIErrorCode.INVALID_CREDENTIALS, status.HTTP_400_BAD_REQUEST)
        try:
            token = RefreshToken(refresh_token)
            token.blacklist()
        except Exception:
            pass  # Token may already be blacklisted or expired — still log out
        logout(request)
        return Response(status=status.HTTP_205_RESET_CONTENT)

# Location-based Attendance View

class SubmitLocationView(generics.GenericAPIView):
    serializer_class = SubmitLocationSerializer
    permission_classes = [IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = 'burst'

    @extend_schema(
        request=SubmitLocationSerializer,
        responses={
            200: OpenApiResponse(response=AttendanceMarkedSerializer, description='Attendance marked successfully.'),
            400: OpenApiResponse(response=APIErrorSerializer, description='Token/coordinates/session validation error.'),
            429: OpenApiResponse(response=APIErrorSerializer, description='Burst rate limit exceeded.'),
        },
        examples=[
            OpenApiExample(
                'Location out of range',
                value={'error': 'Location is out of range', 'code': 'location_out_of_range'},
                response_only=True,
                status_codes=['400'],
            ),
        ],
    )
    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return api_error('Invalid latitude, longitude, or attendance_token.', APIErrorCode.MISSING_REQUIRED_FIELDS, status.HTTP_400_BAD_REQUEST, serializer.errors)
        
        lat_float = serializer.validated_data['latitude']
        lon_float = serializer.validated_data['longitude']
        attendance_token = serializer.validated_data['attendance_token']

        try:
            token = AttendanceToken.objects.get(token=attendance_token, is_active=True)
        except AttendanceToken.DoesNotExist:
            return api_error('Invalid or expired token', APIErrorCode.INVALID_OR_EXPIRED_TOKEN, status.HTTP_400_BAD_REQUEST)

        attendance = Attendance.objects.filter(course=token.course, date=timezone.localdate()).first()

        if not attendance or not attendance.is_session_valid:
            return api_error('This attendance session has expired.', APIErrorCode.SESSION_EXPIRED, status.HTTP_400_BAD_REQUEST)

        if attendance.is_within_radius(lat_float, lon_float):
            user = request.user
            if hasattr(user, 'student'):
                student = user.student
                if token.course.students.filter(pk=student.pk).exists():
                    from django.db import transaction
                    
                    # Add student to attendance with location coordinates
                    with transaction.atomic():
                        locked_attendance = Attendance.objects.select_for_update().get(pk=attendance.pk)
                        AttendanceStudent.objects.get_or_create(
                            attendance=locked_attendance,
                            student=student,
                            defaults={
                                'latitude': lat_float,
                                'longitude': lon_float
                            }
                        )
                        # CRITICAL FIX: Add student to present_students M2M field
                        locked_attendance.present_students.add(student)
                        
                    return Response({'status': 'Attendance marked successfully'}, status=status.HTTP_200_OK)
            return api_error('Student not enrolled in this course', APIErrorCode.STUDENT_NOT_ENROLLED, status.HTTP_400_BAD_REQUEST)

        # Calculate exact distance using geopy
        from geopy.distance import geodesic
        lecturer_location = (attendance.lecturer_latitude, attendance.lecturer_longitude)
        student_location = (lat_float, lon_float)
        distance_km = geodesic(lecturer_location, student_location).kilometers
        distance_meters = distance_km * 1000
        
        return api_error(
            f'Location is out of range. Distance: {distance_meters:.2f}m '
            f'(Max 3000m)', 
            APIErrorCode.LOCATION_OUT_OF_RANGE, 
            status.HTTP_400_BAD_REQUEST
        )

# Student Attendance History View

class StudentAttendanceHistoryView(generics.GenericAPIView):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Fetch the current user and the corresponding student object
        user = self.request.user
        student = get_object_or_404(Student, user=user)

        # Retrieve attendance records where the student was present
        attendance_records = Attendance.objects.filter(
            present_students=student
        ).select_related('course').order_by('-date')

        # Categorize records by course code and order by date descending within each course
        categorized_records = defaultdict(list)
        for attendance in attendance_records:
            course_code = attendance.course.course_code
            categorized_records[course_code].append({
                'date': attendance.date.strftime('%Y-%m-%d'),
            })

        # Prepare the response data
        response_data = [{'course_code': course, 'attendances': records} for course, records in categorized_records.items()]

        return Response(response_data)
    
#Lecturer Attendance History View
class LecturerAttendanceHistoryView(generics.GenericAPIView):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Fetch the current user and the corresponding lecturer object
        user = self.request.user
        lecturer = get_object_or_404(Lecturer, user=user)

        # Retrieve attendance records for courses taught by the lecturer
        attendance_records = Attendance.objects.filter(
            course__lecturer=lecturer
        ).select_related('course').order_by('-date')

        # Categorize records by course code and order by date descending within each course
        categorized_records = defaultdict(list)
        for attendance in attendance_records:
            course_code = attendance.course.course_code
            categorized_records[course_code].append({
                'date': attendance.date.strftime('%Y-%m-%d'),
            })

        # Prepare the response data
        response_data = [{'course_code': course, 'attendances': records} for course, records in categorized_records.items()]

        return Response(response_data)
# Lecturer Location View
class LecturerLocationView(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = SubmitLocationSerializer  # For drf_spectacular schema

    @extend_schema(
        request=LecturerLocationRequestSerializer,
        responses={
            200: OpenApiResponse(response=LecturerLocationResponseSerializer, description='Lecturer coordinates lookup success.'),
            400: OpenApiResponse(response=APIErrorSerializer, description='Token/coordinate availability failure.'),
            403: OpenApiResponse(response=APIErrorSerializer, description='Not authorized for requested course.'),
        },
    )
    def post(self, request, *args, **kwargs):
        token_value = request.data.get('token')

        try:
            token = AttendanceToken.objects.get(token=token_value, is_active=True)
            course = token.course
            lecturer = course.lecturer

            # Only enrolled students or the course lecturer should access this
            if not request.user.is_superuser:
                if hasattr(request.user, 'student'):
                    if not course.students.filter(pk=request.user.student.pk).exists():
                        return api_error('Not enrolled in this course.', APIErrorCode.STUDENT_NOT_ENROLLED, status.HTTP_403_FORBIDDEN)
                elif hasattr(request.user, 'lecturer'):
                    if request.user.lecturer != lecturer:
                        return api_error('Not your course.', APIErrorCode.COURSE_FORBIDDEN, status.HTTP_403_FORBIDDEN)
                else:
                    return api_error('Unauthorized.', APIErrorCode.UNAUTHORIZED, status.HTTP_403_FORBIDDEN)

            # Get active attendance session for today
            attendance = Attendance.objects.filter(
                course=course,
                date=timezone.localdate(),
                is_active=True
            ).first()

            if not attendance:
                return api_error('No active attendance session for this course today.', APIErrorCode.SESSION_EXPIRED, status.HTTP_400_BAD_REQUEST)

            if attendance.lecturer_latitude is None or attendance.lecturer_longitude is None:
                return api_error('Session coordinates not set.', APIErrorCode.LECTURER_COORDINATES_NOT_SET, status.HTTP_400_BAD_REQUEST)

            return Response({
                'longitude': attendance.lecturer_longitude,
                'latitude': attendance.lecturer_latitude,
                'token': token.token
            }, status=status.HTTP_200_OK)

        except AttendanceToken.DoesNotExist:
            return api_error('Invalid or expired token.', APIErrorCode.INVALID_OR_EXPIRED_TOKEN, status.HTTP_400_BAD_REQUEST)
